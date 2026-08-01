from pathlib import Path

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import seaborn as sns

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
OUTPUT_DIR = REPO_ROOT / "outputs" / "figures"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

score_columns = [f"Q{i}" for i in range(1, 21)]
year_folders = [
    DATA_DIR / "2022_23",
    DATA_DIR / "2023_24",
    DATA_DIR / "2024_25",
]


def load_sheet_as_dataframe(file_path: Path) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return pd.DataFrame()

    header = rows[0]
    data = rows[1:]
    return pd.DataFrame(data, columns=header)


records = []
for folder in year_folders:
    if not folder.exists():
        continue

    for file in sorted(folder.glob("*.xlsx")):
        df = load_sheet_as_dataframe(file)
        file_name = file.stem

        missing_columns = [col for col in ["Gender", *score_columns, "Unique Identifier"] if col not in df.columns]
        if missing_columns:
            print(f"Skipping {file.name}: missing columns {missing_columns}")
            continue

        year_label = folder.name
        grade = file_name.split("Grade_")[1].split("_")[0]

        df = df.copy()
        df["year"] = year_label
        df["grade"] = f"Grade {grade}"
        df[score_columns] = df[score_columns].apply(pd.to_numeric, errors="coerce").fillna(0)
        df["total_score"] = df[score_columns].sum(axis=1)
        records.append(df)


if not records:
    raise FileNotFoundError("No Excel files were found in the year-wise folders.")

combined = pd.concat(records, ignore_index=True)
combined["Gender"] = combined["Gender"].fillna("Unknown").astype(str).str.strip().str.lower()

# Basic year-wise insights.
year_summary = (
    combined.groupby("year")
    .agg(
        students=("Unique Identifier", "count"),
        average_score=("total_score", "mean"),
    )
    .reset_index()
)

# Calculate a gender split per year.
gender_counts = (
    combined.groupby(["year", "Gender"])
    .size()
    .unstack(fill_value=0)
    .reset_index()
)

# Question-wise mean performance across years.
question_year_summary = combined.groupby("year")[score_columns].mean().reset_index()

# Print a clean summary in the terminal.
print("Year-wise analysis summary")
print("=" * 80)
print(year_summary.to_string(index=False))
print("\nGender distribution by year")
print("=" * 80)
print(gender_counts.to_string(index=False))
print("\nAverage question performance by year")
print("=" * 80)
print(question_year_summary.to_string(index=False))

# Plot 1: average score by year.
plt.figure(figsize=(8, 5))
sns.barplot(data=year_summary, x="year", y="average_score", palette="viridis")
plt.title("Average Score by Year")
plt.xlabel("Year")
plt.ylabel("Average Marks")
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "average_score_by_year.png", dpi=200)
plt.close()

# Plot 2: question-wise mean by year.
plt.figure(figsize=(12, 6))
for year in question_year_summary["year"]:
    year_values = question_year_summary.loc[question_year_summary["year"] == year, score_columns].iloc[0]
    plt.plot(score_columns, year_values, marker="o", label=year)

plt.title("Average Question Performance by Year")
plt.xlabel("Question")
plt.ylabel("Average Correct Response")
plt.xticks(rotation=45)
plt.legend(title="Year")
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "question_performance_by_year.png", dpi=200)
plt.close()

# Plot 3: gender distribution by year.
plt.figure(figsize=(9, 5))
for gender in ["male", "female", "unknown"]:
    if gender in gender_counts.columns:
        plt.plot(gender_counts["year"], gender_counts[gender], marker="o", label=gender)

plt.title("Gender Distribution by Year")
plt.xlabel("Year")
plt.ylabel("Count")
plt.legend(title="Gender")
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "gender_distribution_by_year.png", dpi=200)
plt.close()

print(f"\nCharts saved to: {OUTPUT_DIR}")
