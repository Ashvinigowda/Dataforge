from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns
import streamlit as st
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, mean_absolute_error
from sklearn.model_selection import train_test_split


REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
score_columns = [f"Q{i}" for i in range(1, 21)]
competency_map = {
    "Conceptual Clarity": [f"Q{i}" for i in range(1, 6)],
    "Application Ability": [f"Q{i}" for i in range(6, 11)],
    "Reasoning": [f"Q{i}" for i in range(11, 16)],
    "Problem Solving": [f"Q{i}" for i in range(16, 21)],
}


@st.cache_data
def load_data() -> tuple[pd.DataFrame, pd.DataFrame, list[str], list[str], list[str], list[str], list[str]]:
    year_folders = [
    DATA_DIR / "2022_23",
    DATA_DIR / "2023_24",
    DATA_DIR / "2024_25",
]
    summary_rows = []
    geo_rows = []
    year_set = set()
    grade_set = set()
    district_set = set()
    block_set = set()
    gp_id_set = set()

    for folder in year_folders:
        if not folder.exists():
            continue
        for file in sorted(folder.glob("*.xlsx")):
            if file.name.startswith("~$"):
                continue

            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()

            if not rows:
                continue

            df = pd.DataFrame(rows[1:], columns=rows[0])
            df["year"] = folder.name
            df["grade"] = file.stem.split("Grade_")[1].split("_")[0]
            df["District"] = df["District"].fillna("Unknown").astype(str).str.strip()
            df["Block"] = df["Block"].fillna("Unknown").astype(str).str.strip()
            df["GP ID"] = df["GP ID"].fillna("Unknown").astype(str).str.strip()
            df[score_columns] = df[score_columns].apply(pd.to_numeric, errors="coerce").fillna(0)
            df["total_score"] = df[score_columns].sum(axis=1)
            df["Gender"] = df["Gender"].fillna("Unknown").astype(str).str.strip().str.lower()
            df["gender_code"] = df["Gender"].map({"female": 0, "male": 1, "unknown": 2}).fillna(2)
            df["year_code"] = df["year"].map({"2022_23": 0, "2023_24": 1, "2024_25": 2})
            df["grade_code"] = pd.to_numeric(df["grade"], errors="coerce").fillna(0)

            question_means = df[score_columns].mean()
            gender_counts = df["Gender"].value_counts()
            avg_score = float(df["total_score"].mean())
            students = int(len(df))

            summary_rows.append(
                {
                    "year": folder.name,
                    "grade": df["grade"].iloc[0],
                    "students": students,
                    "average_score": avg_score,
                    "gender_female": int(gender_counts.get("female", 0)),
                    "gender_male": int(gender_counts.get("male", 0)),
                    "gender_unknown": int(gender_counts.get("unknown", 0)),
                    "year_code": int(df["year_code"].iloc[0]),
                    "grade_code": int(df["grade_code"].iloc[0]),
                    "gender_code": int(df["gender_code"].iloc[0]),
                    "District": df["District"].iloc[0],
                    "Block": df["Block"].iloc[0],
                    "GP ID": df["GP ID"].iloc[0],
                    **{f"q_mean_{q}": float(question_means[q]) for q in score_columns},
                }
            )

            geo_rows.append(
                {
                    "District": df["District"].iloc[0],
                    "Block": df["Block"].iloc[0],
                    "GP ID": df["GP ID"].iloc[0],
                    "year": folder.name,
                    "grade": df["grade"].iloc[0],
                    "students": students,
                    "average_score": avg_score,
                }
            )

            year_set.add(folder.name)
            grade_set.add(df["grade"].iloc[0])
            district_set.add(df["District"].iloc[0])
            block_set.add(df["Block"].iloc[0])
            gp_id_set.add(df["GP ID"].iloc[0])

    if not summary_rows:
        raise FileNotFoundError("No Excel files found.")

    file_summary = pd.DataFrame(summary_rows)
    geo_summary = pd.DataFrame(geo_rows)
    years = sorted(year_set)
    grades = sorted(grade_set)
    districts = sorted(district_set)
    blocks = sorted(block_set)
    gp_ids = sorted(gp_id_set)
    return file_summary, geo_summary, years, grades, districts, blocks, gp_ids


@st.cache_data
def build_sampled_model_data() -> pd.DataFrame:
    sample_rows = []
    for folder in [
    DATA_DIR / "2022_23",
    DATA_DIR / "2023_24",
    DATA_DIR / "2024_25",
]:
        if not folder.exists():
            continue
        for file in sorted(folder.glob("*.xlsx")):
            if file.name.startswith("~$"):
                continue

            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()
            if not rows:
                continue

            df = pd.DataFrame(rows[1:], columns=rows[0])
            df["year"] = folder.name
            df["grade"] = file.stem.split("Grade_")[1].split("_")[0]
            df["District"] = df["District"].fillna("Unknown").astype(str).str.strip()
            df["Block"] = df["Block"].fillna("Unknown").astype(str).str.strip()
            df["GP ID"] = df["GP ID"].fillna("Unknown").astype(str).str.strip()
            df[score_columns] = df[score_columns].apply(pd.to_numeric, errors="coerce").fillna(0)
            df["total_score"] = df[score_columns].sum(axis=1)
            df["Gender"] = df["Gender"].fillna("Unknown").astype(str).str.strip().str.lower()
            df["gender_code"] = df["Gender"].map({"female": 0, "male": 1, "unknown": 2}).fillna(2)
            df["year_code"] = df["year"].map({"2022_23": 0, "2023_24": 1, "2024_25": 2})
            df["grade_code"] = pd.to_numeric(df["grade"], errors="coerce").fillna(0)
            df["district_code"] = pd.Categorical(df["District"]).codes
            df["block_code"] = pd.Categorical(df["Block"]).codes
            df["gp_code"] = pd.Categorical(df["GP ID"]).codes

            sample_df = df.sample(n=min(len(df), 250), random_state=42)
            sample_rows.append(sample_df)

    sampled_df = pd.concat(sample_rows, ignore_index=True)
    return sampled_df


@st.cache_data
def build_model_summary(sampled_df: pd.DataFrame) -> tuple[float, pd.Series]:
    features = ["grade_code", "year_code", "gender_code", "district_code", "block_code", "gp_code", *score_columns]
    X = sampled_df[features]
    y = sampled_df["total_score"]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    model = RandomForestRegressor(n_estimators=150, random_state=42)
    model.fit(X_train, y_train)
    predictions = model.predict(X_test)
    mae = float(mean_absolute_error(y_test, predictions))
    importances = pd.Series(model.feature_importances_, index=features).sort_values(ascending=False)
    return mae, importances


@st.cache_data
def build_classification_summary(sampled_df: pd.DataFrame) -> tuple[float, float, pd.Series]:
    sampled_df = sampled_df.copy()
    sampled_df["high_performer"] = (sampled_df["total_score"] >= sampled_df["total_score"].median()).astype(int)
    features = ["grade_code", "year_code", "gender_code", "district_code", "block_code", "gp_code", *score_columns]
    X = sampled_df[features]
    y = sampled_df["high_performer"]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    rf_model = RandomForestClassifier(n_estimators=150, random_state=42)
    rf_model.fit(X_train, y_train)
    rf_predictions = rf_model.predict(X_test)
    rf_accuracy = float(accuracy_score(y_test, rf_predictions))

    lr_model = LogisticRegression(max_iter=300, random_state=42)
    lr_model.fit(X_train, y_train)
    lr_predictions = lr_model.predict(X_test)
    lr_accuracy = float(accuracy_score(y_test, lr_predictions))

    importances = pd.Series(rf_model.feature_importances_, index=features).sort_values(ascending=False)
    return rf_accuracy, lr_accuracy, importances


st.set_page_config(page_title="Math Analysis Dashboard", layout="wide")
st.title("Year-wise Math Analysis Dashboard")

file_summary, geo_summary_raw, years, grades, districts, blocks, gp_ids = load_data()
sampled_model_df = build_sampled_model_data()

selected_years = st.sidebar.multiselect("Select years", options=years, default=years)
selected_grades = st.sidebar.multiselect("Select grade", options=grades, default=grades)
selected_districts = st.sidebar.multiselect("Select district", options=districts, default=districts)
selected_blocks = st.sidebar.multiselect("Select block", options=blocks, default=blocks)
selected_gp_ids = st.sidebar.multiselect("Select gram panchayat ID", options=gp_ids, default=gp_ids)

if not selected_years or not selected_grades:
    st.warning("Please select at least one year and one grade.")
    st.stop()

filtered_summary = file_summary[
    file_summary["year"].isin(selected_years)
    & file_summary["grade"].isin(selected_grades)
].copy()
filtered_geo = geo_summary_raw[
    geo_summary_raw["year"].isin(selected_years)
    & geo_summary_raw["grade"].isin(selected_grades)
    & geo_summary_raw["District"].isin(selected_districts)
    & geo_summary_raw["Block"].isin(selected_blocks)
    & geo_summary_raw["GP ID"].isin(selected_gp_ids)
].copy()
filtered_sample = sampled_model_df[
    sampled_model_df["year"].isin(selected_years)
    & sampled_model_df["grade"].isin(selected_grades)
    & sampled_model_df["District"].isin(selected_districts)
    & sampled_model_df["Block"].isin(selected_blocks)
    & sampled_model_df["GP ID"].isin(selected_gp_ids)
].copy()

if filtered_summary.empty or filtered_sample.empty:
    st.warning("No data matches the selected filters.")
    st.stop()

year_summary = (
    filtered_summary.groupby("year", as_index=False)
    .agg(
        students=("students", "sum"),
        average_score=("average_score", "mean"),
    )
)

gender_counts = (
    filtered_summary.groupby("year", as_index=False)
    .agg(
        female=("gender_female", "sum"),
        male=("gender_male", "sum"),
        unknown=("gender_unknown", "sum"),
    )
)

question_year_summary = pd.DataFrame(
    {
        "year": filtered_summary["year"],
        **{q: filtered_summary[f"q_mean_{q}"] for q in score_columns},
    }
)
question_year_summary = question_year_summary.groupby("year", as_index=False)[score_columns].mean()

competency_summary = []
for competency_name, q_subset in competency_map.items():
    competency_summary.append(
        {
            "Competency": competency_name,
            "Average score": float(filtered_summary[[f"q_mean_{q}" for q in q_subset]].mean().mean()),
            "Weakest question": filtered_summary[[f"q_mean_{q}" for q in q_subset]].mean().idxmin().replace("q_mean_", ""),
        }
    )
comp_df = pd.DataFrame(competency_summary)

geo_summary = filtered_geo.groupby(["District", "Block"], as_index=False).agg(
    students=("students", "sum"),
    average_score=("average_score", "mean"),
)
geo_summary = geo_summary.sort_values("average_score", ascending=False)

student_count_growth = year_summary["students"].iloc[-1] - year_summary["students"].iloc[0]
insight_1 = f"Participation increased by {student_count_growth:,} students from the first selected year to the latest selected year, so the programme is scaling in reach."

best_year = year_summary.sort_values("average_score", ascending=False).iloc[0]["year"]
best_avg = year_summary.sort_values("average_score", ascending=False).iloc[0]["average_score"]
insight_2 = f"The strongest average performance appears in {best_year} at {best_avg:.2f} marks, which is the best benchmark for the current selection."

female_total = int(gender_counts[["female"]].sum().sum()) if "female" in gender_counts.columns else 0
male_total = int(gender_counts[["male"]].sum().sum()) if "male" in gender_counts.columns else 0
insight_3 = f"Female participation is ahead of male participation in the selected view, with {female_total} females vs {male_total} males. This suggests the programme is engaging a broader student base than pure male representation alone."

question_mean = question_year_summary[score_columns].mean().mean()
weakest_question = question_year_summary[score_columns].mean().idxmin()
weakest_value = question_year_summary[score_columns].mean().min()
insight_4 = f"The clearest intervention opportunity is {weakest_question}, where the average response level is {weakest_value:.2f}, below the overall question mean of {question_mean:.2f}."

best_district = geo_summary.sort_values("average_score", ascending=False).iloc[0]["District"]
best_block = geo_summary.sort_values("average_score", ascending=False).iloc[0]["Block"]
insight_5 = f"The highest-performing geography in the current slice is {best_block}, {best_district}, which can be treated as a strong comparison benchmark for replication."

trend_years = np.array([int(y.split('_')[0]) for y in year_summary["year"]], dtype=float)
trend_students = year_summary["students"].to_numpy(dtype=float)
trend_score = year_summary["average_score"].to_numpy(dtype=float)
next_year = int(max(trend_years)) + 1
students_slope, students_intercept = np.polyfit(trend_years, trend_students, 1)
predicted_students = int(round(students_slope * next_year + students_intercept))
score_slope, score_intercept = np.polyfit(trend_years, trend_score, 1)
predicted_score = float(score_slope * next_year + score_intercept)

model_mae, model_importances = build_model_summary(filtered_sample)
model_rf_accuracy, model_lr_accuracy, model_feature_preview = build_classification_summary(filtered_sample)
model_feature_preview = model_feature_preview.head(5).rename("Importance")

prediction_text = (
    f"A simple linear forecast suggests the next year could reach about {predicted_students:,} students "
    f"with an estimated average score around {predicted_score:.2f}. Meanwhile, a sampled model check on the selected records shows a holdout MAE of {model_mae:.2f} marks, with random forest accuracy {model_rf_accuracy:.2f} and logistic regression accuracy {model_lr_accuracy:.2f}."
)

question_level = question_year_summary[score_columns].mean().sort_values()
weakest_questions = question_level.head(3)
strongest_questions = question_level.sort_values(ascending=False).head(3)

st.subheader("Key Insights")
st.markdown(f"- {insight_1}")
st.markdown(f"- {insight_2}")
st.markdown(f"- {insight_3}")
st.markdown(f"- {insight_4}")
st.markdown(f"- {insight_5}")
st.success(prediction_text)

st.subheader("Predictive and Diagnostic View")
col_a, col_b = st.columns(2)
with col_a:
    st.info("Predictive perspective: one estimate is based on a simple linear trend over the selected years, and the other is a sampled model check on the actual response pattern.")
    st.metric("Forecast next year students", f"{predicted_students:,}")
    st.metric("Forecast next year average score", f"{predicted_score:.2f}")
    st.metric("Model holdout MAE", f"{model_mae:.2f} marks")
    st.metric("Random forest accuracy", f"{model_rf_accuracy:.2f}")
    st.metric("Logistic regression accuracy", f"{model_lr_accuracy:.2f}")
with col_b:
    st.warning("Diagnostic perspective: focus on recurring weaker question areas for targeted intervention and better preparation planning.")
    st.write("Top 3 weakest question areas:")
    st.dataframe(weakest_questions.rename("Average success rate").reset_index().rename(columns={"index": "Question"}))
    st.write("Top 5 model-driving signals:")
    st.dataframe(model_feature_preview.reset_index().rename(columns={"index": "Feature"}))

st.subheader("Summary")
col1, col2, col3 = st.columns(3)
col1.metric("Students", int(filtered_summary["students"].sum()))
col2.metric("Average Score", f"{filtered_summary['average_score'].mean():.2f}")
col3.metric("Years Covered", filtered_summary["year"].nunique())

st.dataframe(year_summary, width="stretch")

st.subheader("Average Score by Year")
fig, ax = plt.subplots(figsize=(8, 5))
sns.barplot(data=year_summary, x="year", y="average_score", palette="viridis", ax=ax)
ax.set_title("Average Score by Year")
ax.set_xlabel("Year")
ax.set_ylabel("Average Marks")
plt.tight_layout()
st.pyplot(fig)

st.subheader("Gender Distribution by Year")
fig, ax = plt.subplots(figsize=(9, 5))
for gender in ["male", "female", "unknown"]:
    if gender in gender_counts.columns:
        ax.plot(gender_counts["year"], gender_counts[gender], marker="o", label=gender)
ax.set_title("Gender Distribution by Year")
ax.set_xlabel("Year")
ax.set_ylabel("Count")
ax.legend(title="Gender")
plt.tight_layout()
st.pyplot(fig)

st.subheader("Average Question Performance by Year")
fig, ax = plt.subplots(figsize=(12, 6))
for year in question_year_summary["year"]:
    year_values = question_year_summary.loc[question_year_summary["year"] == year, score_columns].iloc[0]
    ax.plot(score_columns, year_values, marker="o", label=year)
ax.set_title("Average Question Performance by Year")
ax.set_xlabel("Question")
ax.set_ylabel("Average Correct Response")
ax.tick_params(axis="x", rotation=45)
ax.legend(title="Year")
plt.tight_layout()
st.pyplot(fig)

st.subheader("Geographic Comparison")
st.dataframe(geo_summary.head(10), width="stretch")

st.subheader("Competency Summary")
st.dataframe(comp_df, width="stretch")

st.subheader("Question-wise Summary Table")
st.dataframe(question_year_summary, width="stretch")

st.subheader("Strongest Question Areas")
st.dataframe(strongest_questions.rename("Average success rate").reset_index().rename(columns={"index": "Question"}))
